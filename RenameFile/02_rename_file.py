import os
import pandas as pd
from openpyxl import load_workbook
import re


def process_video_files(excel_path, folder_path):
    """
    处理视频文件重命名和Excel标记，不区分大小写匹配，重命名用大写

    Args:
        excel_path (str): Excel文件路径
        folder_path (str): 视频文件夹路径
    """

    # 支持的视频格式
    video_extensions = ['.mp4', '.avi', '.mov', '.mkv', '.wmv', '.flv', '.m4v', '.webm', '.MP4', '.AVI', '.MOV', '.MKV', '.mpeg']

    try:
        # 使用openpyxl加载工作簿，保持格式
        wb = load_workbook(excel_path)

        # 检查工作表是否存在
        if 'Auto超链接' not in wb.sheetnames:
            print("Excel中未找到'Auto超链接'工作表")
            return

        ws = wb['Auto超链接']

        # 查找列索引
        title_col_idx = None
        exist_col_idx = None

        # 查找标题列和存在列的位置
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value == '标题':
                title_col_idx = col_idx
            elif cell_value == '存在':
                exist_col_idx = col_idx

        if title_col_idx is None:
            print("Excel中未找到'标题'列")
            return

        # 如果不存在列，则在最后一列后添加
        if exist_col_idx is None:
            exist_col_idx = ws.max_column + 1
            ws.cell(row=1, column=exist_col_idx, value='存在')
            print("已添加'存在'列")

        # 收集标题和对应的行索引（存储原始标题和清理后的小写标题）
        titles_info = {}  # {小写标题: (原始标题, 行索引)}
        existing_titles_lower = set()  # 已经标记为存在的小写标题

        for row_idx in range(2, ws.max_row + 1):
            title_value = ws.cell(row=row_idx, column=title_col_idx).value
            exist_value = ws.cell(row=row_idx, column=exist_col_idx).value

            if title_value and pd.notna(title_value):
                original_title = str(title_value).strip()
                lower_title = original_title.lower()
                titles_info[lower_title] = (original_title, row_idx)
                if exist_value == '已存在':
                    existing_titles_lower.add(lower_title)

        titles_to_process_lower = [title for title in titles_info.keys() if title not in existing_titles_lower]

        print(f"找到 {len(titles_info)} 个标题，其中 {len(titles_to_process_lower)} 个需要处理")

        # 获取文件夹中的所有视频文件
        video_files = []
        for file in os.listdir(folder_path):
            if any(file.lower().endswith(ext.lower()) for ext in video_extensions):
                video_files.append(file)

        print(f"找到 {len(video_files)} 个视频文件")

        # 处理每个视频文件
        processed_count = 0
        for video_file in video_files:
            file_path = os.path.join(folder_path, video_file)
            file_name, file_ext = os.path.splitext(video_file)

            # 按"-"分割文件名
            parts = file_name.split('-')

            if len(parts) >= 2:
                # 构建匹配模式（转换为小写用于匹配）
                if len(parts) == 2:
                    # 两部分的情况
                    match_pattern1 = f"{parts[0]}-{parts[1]}".lower()
                    match_pattern2 = f"{parts[0]}{parts[1]}".lower()
                    serial_number = None
                else:
                    # 三部分或更多的情况
                    match_pattern1 = f"{parts[0]}-{parts[1]}".lower()
                    match_pattern2 = f"{parts[0]}{parts[1]}".lower()
                    serial_number = parts[2]  # 保持原始大小写，后续会转为大写

                # 在需要处理的标题中查找匹配项（使用小写比较）
                matched_lower_title = None
                for lower_title in titles_to_process_lower:
                    # 清理标题和模式中的特殊字符用于比较
                    clean_lower_title = lower_title.replace(" ", "").replace("-", "").replace("_", "")
                    clean_pattern1 = match_pattern1.replace(" ", "").replace("-", "").replace("_", "")
                    clean_pattern2 = match_pattern2.replace(" ", "").replace("-", "").replace("_", "")

                    # 多种匹配方式（不区分大小写）
                    if (match_pattern1 in lower_title or
                            match_pattern2 in lower_title or
                            lower_title.startswith(match_pattern1) or
                            lower_title.startswith(match_pattern2) or
                            clean_pattern1 in clean_lower_title or
                            clean_pattern2 in clean_lower_title or
                            lower_title.endswith(match_pattern1) or
                            lower_title.endswith(match_pattern2)):
                        matched_lower_title = lower_title
                        break

                if not matched_lower_title:
                    # 检查是否已经存在
                    for lower_title in existing_titles_lower:
                        clean_lower_title = lower_title.replace(" ", "").replace("-", "").replace("_", "")
                        clean_pattern1 = match_pattern1.replace(" ", "").replace("-", "").replace("_", "")
                        clean_pattern2 = match_pattern2.replace(" ", "").replace("-", "").replace("_", "")

                        if (match_pattern1 in lower_title or
                                match_pattern2 in lower_title or
                                clean_pattern1 in clean_lower_title or
                                clean_pattern2 in clean_lower_title):
                            original_title = titles_info[lower_title][0]
                            #print(f"文件 {video_file} 匹配到已存在的标题 '{original_title}'，跳过处理")
                            break
                    else:
                        print(f"未找到匹配的标题: {video_file}")
                    continue

                # 获取原始标题和行索引
                original_title, row_idx = titles_info[matched_lower_title]

                # 构建新文件名（全部转为大写）
                if serial_number:
                    # 序号也转为大写
                    serial_number_upper = serial_number.upper()
                    new_file_name = f"{original_title.upper()}-{serial_number_upper}{file_ext}"
                else:
                    new_file_name = f"{original_title.upper()}{file_ext}"

                # 清理文件名中的非法字符
                new_file_name = re.sub(r'[<>:"/\\|?*]', '', new_file_name)

                # 重命名文件
                new_file_path = os.path.join(folder_path, new_file_name)

                try:
                    # 检查目标文件是否已存在
                    if os.path.exists(new_file_path):
                        print(f"目标文件已存在，跳过重命名: {new_file_name}")
                        continue

                    os.rename(file_path, new_file_path)
                    print(f"重命名: {video_file} -> {new_file_name}")

                    # 在Excel中标记为"已存在"
                    ws.cell(row=row_idx, column=exist_col_idx, value='已存在')
                    print(f"标记标题 '{original_title}' 为已存在")

                    # 从待处理列表中移除
                    if matched_lower_title in titles_to_process_lower:
                        titles_to_process_lower.remove(matched_lower_title)

                    processed_count += 1

                except Exception as e:
                    print(f"重命名失败 {video_file}: {e}")

        # 保存Excel文件（保持原有格式）
        try:
            wb.save(excel_path)
            print(f"Excel文件已保存，共处理了 {processed_count} 个文件")
        except Exception as e:
            print(f"保存Excel失败，请确保文件没有被其他程序占用: {e}")

    except Exception as e:
        print(f"处理过程中出错: {e}")


# 使用示例
if __name__ == "__main__":
    # 配置参数
    # 设置你的Excel文件路径和视频文件夹路径
    excel_file_path = r"../DownloadVideo/Resource/BT.xlsx"  # Excel文件路径
    video_folder_path = r"E:\02_ACG\映画\步兵系列"  # 视频文件夹路径

    process_video_files(excel_file_path, video_folder_path)