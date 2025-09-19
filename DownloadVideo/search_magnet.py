import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import re


def extract_magnet_links_from_excel():
    # Excel文件路径
    excel_file = 'BT.xlsx'  # 替换为你的Excel文件路径

    try:
        # 1. 从Excel读取URL
        wb = load_workbook(excel_file)
        sheet2 = wb['Sheet2']
        sheet1 = wb['Sheet1']

        # 2. 确定起始行：从Sheet1中找到最后一个有值的B列单元格
        start_row = 2  # 默认从第2行开始
        row = 2
        while True:
            cell_value = sheet1[f'B{row}'].value
            if cell_value is None or str(cell_value).strip() == '':
                start_row = row  # 找到第一个空单元格
                break
            row += 1
            # 安全限制，防止无限循环
            if row > 1000:
                print("警告：已达到最大行数限制")
                start_row = row
                break

        print(f"检测到Sheet1中B列最后一个有值的行是第 {start_row - 1} 行")
        print(f"将从Sheet2的D{start_row}开始处理")

        # 3. 获取D列从起始行开始的所有URL，直到遇到空单元格
        urls = []
        row = start_row
        while True:
            cell_value = sheet2[f'D{row}'].value
            if cell_value is None or str(cell_value).strip() == '':
                break
            urls.append((row, cell_value))  # 保存行号和URL
            row += 1
            # 安全限制，防止无限循环
            if row > start_row + 1000:
                print("警告：已达到最大处理行数限制")
                break

        if not urls:
            print(f"Sheet2的D列从D{start_row}开始没有找到URL")
            return

        print(f"从Excel获取了 {len(urls)} 个待处理的URL")

        # 4. 设置Chrome浏览器选项
        chrome_options = Options()
        #chrome_options.add_argument('--headless')  # 无头模式，不需要显示浏览器界面
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument(
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')

        # 初始化Chrome浏览器
        service = Service('chromedriver.exe')
        driver = webdriver.Chrome(service=service, options=chrome_options)

        # 处理每个URL
        processed_count = 0
        for row_num, url in urls:
            try:
                print(f"\n处理第 {row_num} 行的URL: {url}")

                # 5. 打开网页
                print("正在打开网页...")
                driver.get(url)
                time.sleep(3)  # 等待页面加载

                # 6. 检查是否有"请点此进入"或类似提示
                try:
                    page_text = driver.page_source
                    if "请点此进入" in page_text or "请点击进入" in page_text or "点击进入" in page_text:
                        print("检测到需要点击进入的提示")

                        # 尝试找到并点击进入链接
                        entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "进入")
                        if not entry_links:
                            entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "点此")
                        if not entry_links:
                            entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "点击")

                        if entry_links:
                            print(f"找到 {len(entry_links)} 个可能的进入链接")
                            entry_links[0].click()
                            print("已点击进入链接")
                            time.sleep(3)  # 等待新页面加载
                        else:
                            print("未找到明确的进入链接，尝试查找按钮")
                            buttons = driver.find_elements(By.TAG_NAME, "button")
                            for button in buttons:
                                if "进入" in button.text or "点此" in button.text or "点击" in button.text:
                                    button.click()
                                    print("已点击进入按钮")
                                    time.sleep(3)
                                    break
                except Exception as e:
                    print(f"检查进入提示时出错: {e}")

                # 7. 获取页面源代码
                page_source = driver.page_source

                # 8. 使用正则表达式查找magnet链接
                magnet_pattern = r'magnet:\?xt=urn:btih:[a-zA-Z0-9]{40}.*?(?=\'|"|&|<|\\s)'
                magnet_links = re.findall(magnet_pattern, page_source)

                if not magnet_links:
                    # 尝试更宽松的匹配模式
                    magnet_pattern = r'magnet:\?xt=[^\'"\s<>]+'
                    magnet_links = re.findall(magnet_pattern, page_source)

                if magnet_links:
                    # 取第一个找到的magnet链接
                    magnet_link = magnet_links[0]
                    print(f"找到magnet链接: {magnet_link}")

                    # 检查是否有重复的磁力链接
                    is_duplicate = False
                    for check_row in range(2, row_num):
                        existing_link = sheet1[f'B{check_row}'].value
                        if existing_link and existing_link == magnet_link:
                            is_duplicate = True
                            print(f"发现重复的磁力链接，位于B{check_row}")
                            break

                    # 9. 将magnet链接写入Sheet1的对应行
                    sheet1[f'B{row_num}'] = magnet_link

                    # 如果是重复链接，在D列标记"重复"
                    if is_duplicate:
                        sheet1[f'D{row_num}'] = "重复"
                        print(f"重复磁力链接，已在D{row_num}单元格标记")
                    else:
                        # 确保D列没有标记（如果是新处理的行）
                        if sheet1[f'D{row_num}'].value == "重复":
                            sheet1[f'D{row_num}'] = None

                    print(f"magnet链接已写入Sheet1的B{row_num}单元格")
                    processed_count += 1

                    # 每处理5个URL保存一次，防止数据丢失
                    if processed_count % 5 == 0:
                        wb.save(excel_file)
                        print(f"已处理 {processed_count} 个URL，自动保存Excel文件")
                else:
                    print("在页面中没有找到magnet链接")
                    # 保存页面源代码以供调试
                    with open(f"NoFind/debug_page_{row_num}.html", "w", encoding="utf-8") as f:
                        f.write(page_source)
                    print(f"已保存页面源代码到debug_page_{row_num}.html以供调试")

            except Exception as e:
                print(f"处理URL {url} 时发生错误: {e}")
                # 继续处理下一个URL
                continue

        # 处理完成后，再次检查所有B列值，确保标记所有重复项
        print("\n处理完成，开始全面检查重复项...")
        magnet_values = {}
        for row in range(2, sheet1.max_row + 1):
            magnet_value = sheet1[f'B{row}'].value
            if magnet_value and str(magnet_value).strip() != '':
                if magnet_value in magnet_values:
                    # 标记当前行和之前的所有相同行为重复
                    sheet1[f'D{row}'] = "重复"
                    sheet1[f'D{magnet_values[magnet_value]}'] = "重复"
                    print(f"标记重复: 行 {magnet_values[magnet_value]} 和 行 {row}")
                else:
                    magnet_values[magnet_value] = row

        # 最终保存Excel文件
        wb.save(excel_file)
        print(f"\n所有URL处理完成，共处理了 {processed_count} 个URL，Excel文件已保存")

    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")

    finally:
        # 关闭浏览器
        if 'driver' in locals():
            driver.quit()


# 运行函数
if __name__ == "__main__":
    extract_magnet_links_from_excel()
