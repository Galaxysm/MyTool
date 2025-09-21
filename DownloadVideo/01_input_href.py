import os
import time
from datetime import datetime

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


class MagnetLinkExtractor:
    def __init__(self):
        self.driver = None
        self.setup_driver()

    def setup_driver(self):
        """设置Chrome浏览器驱动"""
        chrome_options = Options()
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument(
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

        # 可选：无头模式（取消注释以在后台运行）
        # chrome_options.add_argument('--headless')

        # 设置下载行为（避免弹窗）
        prefs = {
            "download.default_directory": os.getcwd(),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        try:
            service = Service('Resource/chromedriver.exe')
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.implicitly_wait(10)
            print("浏览器驱动初始化成功")
        except Exception as e:
            print(f"浏览器驱动初始化失败: {e}")
            raise

    def click_entry_if_exists(self):
        """检查并点击'请点此进入'等入口链接"""
        entry_phrases = ["请点此进入", "点击进入", "点此进入", "进入", "进入网站", "继续访问", "访问", "进入主页"]

        for phrase in entry_phrases:
            try:
                # 尝试多种定位方式
                selectors = [
                    f"//a[contains(text(), '{phrase}')]",
                    f"//button[contains(text(), '{phrase}')]",
                    f"//div[contains(text(), '{phrase}')]",
                    f"//span[contains(text(), '{phrase}')]",
                    f"//*[contains(@onclick, '{phrase}')]"
                ]

                for selector in selectors:
                    try:
                        element = WebDriverWait(self.driver, 3).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        if element:
                            print(f"找到'{phrase}'入口，正在点击...")
                            element.click()
                            time.sleep(2)  # 等待页面跳转
                            return True
                    except:
                        continue
            except Exception as e:
                continue

        print("未找到需要点击的入口链接")
        return False

    def find_thread_elements(self):
        """在页面中寻找帖子链接和标题"""
        try:
            # 等待页面主体加载完成
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            # 尝试多种可能的选择器来查找帖子链接
            selectors = [
                "//a[contains(@class, 's xst')]",
                "//a[contains(@class, 'xst')]",
                "//a[contains(@id, 'thread_')]",
                "//tbody[contains(@id, 'normalthread_')]//a[@class='s xst']",
                "//*[@id='threadlist']//a[contains(@class, 's xst')]",
                "//*[contains(@id, 'threadlisttableid')]//a[contains(@class, 's xst')]"
            ]

            thread_data = []  # 存储链接和标题的元组列表
            found_elements = False

            for selector in selectors:
                try:
                    # 等待元素出现
                    WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, selector))
                    )

                    links = self.driver.find_elements(By.XPATH, selector)
                    if links:
                        print(f"使用选择器 '{selector}' 找到 {len(links)} 个链接")
                        found_elements = True

                        for link in links:
                            try:
                                href = link.get_attribute('href')
                                text = link.text
                                if href and (href, text) not in thread_data:
                                    thread_data.append((href, text))
                                    print(f"找到链接: {href}, 标题: {text}")
                            except StaleElementReferenceException:
                                print("元素已过期，跳过")
                                continue

                        # 如果找到链接，不再尝试其他选择器
                        if thread_data:
                            break

                except (TimeoutException, NoSuchElementException):
                    print(f"选择器 '{selector}' 未找到元素")
                    continue

            if not found_elements:
                print("尝试备用方法：查找所有包含thread的链接")
                # 备用方法：查找所有包含thread的链接
                all_links = self.driver.find_elements(By.TAG_NAME, "a")
                for link in all_links:
                    try:
                        href = link.get_attribute('href')
                        text = link.text
                        if href and 'thread' in href and 'forum' in href and (href, text) not in thread_data:
                            thread_data.append((href, text))
                    except StaleElementReferenceException:
                        continue

            return thread_data

        except Exception as e:
            print(f"查找元素时出错: {e}")
            return []

    def process_all_links(self, main_url):
        """处理所有链接并提取超链接和标题"""
        try:
            # 访问主页面
            print(f"正在访问主页面: {main_url}")
            self.driver.get(main_url)

            # 等待页面加载
            time.sleep(3)

            # 检查并点击入口链接
            self.click_entry_if_exists()

            # 等待页面加载
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            time.sleep(2)

            # 查找目标元素
            elements = self.find_thread_elements()
            if not elements:
                print("未找到超链接")
                return []

            # 提取所有链接和标题
            print(f"共找到 {len(elements)} 个链接需要处理")
            return elements

        except Exception as e:
            print(f"处理过程中出错: {e}")
            return []

    def get_existing_links(self, excel_file, sheet_name):
        """从Excel文件中获取已存在的所有链接"""
        existing_links = set()

        try:
            if os.path.exists(excel_file):
                # 使用openpyxl直接读取B列数据
                wb = load_workbook(excel_file)

                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    # 遍历B列的所有行
                    for row in range(1, sheet.max_row + 1):
                        cell_value = sheet[f'B{row}'].value
                        if cell_value and isinstance(cell_value, str):
                            existing_links.add(cell_value.strip())

                    print(f"从Excel中读取到 {len(existing_links)} 个已存在的链接")
                else:
                    print(f"工作表 '{sheet_name}' 不存在")
            else:
                print("Excel文件不存在，将创建新文件")
        except Exception as e:
            print(f"读取现有链接时出错: {e}")

        return existing_links

    def save_to_excel(self, thread_data, excel_file, sheet_name):
        """将超链接和标题保存到Excel，如果B列已有数据，则在下方追加，并避免重复"""
        try:
            # 获取已存在的链接
            existing_links = self.get_existing_links(excel_file, sheet_name)

            # 过滤掉已存在的链接
            new_data = [(link, title) for link, title in thread_data if link not in existing_links]

            if not new_data:
                print("没有新的链接需要添加")
                return

            print(f"过滤后，有 {len(new_data)} 个新链接需要保存")

            # 检查Excel文件是否存在
            if os.path.exists(excel_file):
                wb = load_workbook(excel_file)

                # 检查工作表是否存在
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb.create_sheet(sheet_name)
            else:
                wb = load_workbook()
                sheet = wb.create_sheet(sheet_name)

            # 找到B列最后一个非空行
            last_row = 1
            for row in range(1, sheet.max_row + 1):
                if sheet[f'B{row}'].value is not None:
                    last_row = row + 1

            # 写入新链接和标题
            for i, (link, title) in enumerate(new_data, start=last_row):
                sheet[f'B{i}'] = link
                sheet[f'C{i}'] = title  # 将标题写入C列
                sheet[f'D{i}'] = datetime.now()  # 将时间写入D列

            # 保存文件
            wb.save(excel_file)
            print(
                f"成功保存 {len(new_data)} 个新超链接到 {excel_file} 的 {sheet_name} 工作表，从第 {last_row} 行开始")

        except Exception as e:
            print(f"保存到Excel时出错: {e}")

    def close(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()
            print("浏览器已关闭")


def main():
    # 配置参数
    target_url = "https://309r.0i284.net/forum-151-{}.html"
    output_excel = "Resource/BT.xlsx"
    sheet_name = "Auto超链接"
    page = 1
    max_pages = 78

    while page <= max_pages:
        print(f"\n{'=' * 50}")
        print(f"开始处理第 {page}/{max_pages} 页")
        print(f"{'=' * 50}")

        # 创建提取器实例
        extractor = MagnetLinkExtractor()

        try:
            # 提取超链接和标题
            #url = target_url + str(page)               # 拼接形式1
            url = target_url.format(str(page))          # 拼接形式2

            thread_data = extractor.process_all_links(url)

            if thread_data:
                print(f"第 {page} 页找到 {len(thread_data)} 个超链接")
                # 保存到Excel
                extractor.save_to_excel(thread_data, output_excel, sheet_name)
            else:
                print(f"第 {page} 页未找到任何超链接")

                # 如果连续两页都没有找到链接，停止爬取
                if page > 1 and not thread_data:
                    print("连续两页没有找到链接，停止爬取")
                    break

            page += 1
            time.sleep(3)  # 添加延迟避免请求过于频繁

        except Exception as e:
            print(f"程序执行出错: {e}")
        finally:
            # 关闭浏览器
            extractor.close()


if __name__ == "__main__":
    main()