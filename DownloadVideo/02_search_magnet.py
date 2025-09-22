import re
import time
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By


def extract_magnet_links_from_excel():
    # Excel文件路径
    excel_file = 'Resource/BT.xlsx'  # 替换为你的Excel文件路径

    try:
        # 1. 加载Excel文件
        wb = load_workbook(excel_file)
        sheet = wb['Sheet1']  # 只使用Sheet1或Sheet2

        # 2. 找到C列第一个空单元格及其对应的B列URL
        urls_to_process = []
        row = 2  # 从第2行开始（假设第1行是标题）

        while True:
            # 检查C列是否为空
            c_cell_value = sheet[f'C{row}'].value
            b_cell_value = sheet[f'B{row}'].value

            # 如果C列为空且B列有URL，则加入处理列表
            if (c_cell_value is None or str(c_cell_value).strip() == '') and b_cell_value and str(
                    b_cell_value).strip() != '':
                urls_to_process.append((row, b_cell_value))
                print(f"找到待处理的行 {row}: {b_cell_value}")

            # 如果B列为空，说明没有更多URL了，停止搜索
            if b_cell_value is None or str(b_cell_value).strip() == '':
                break

            row += 1
            # 安全限制，防止无限循环
            if row > 10000:
                print("警告：已达到最大行数限制")
                break

        if not urls_to_process:
            print("没有找到需要处理的URL（C列为空且B列有URL的行）")
            return

        print(f"找到 {len(urls_to_process)} 个需要处理的URL")

        # 3. 设置Chrome浏览器选项
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # 无头模式，不需要显示浏览器界面
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument(
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')

        # 初始化Chrome浏览器
        service = Service('Resource/chromedriver.exe')
        driver = webdriver.Chrome(service=service, options=chrome_options)

        # 处理每个URL
        processed_count = 0
        for row_num, url in urls_to_process:
            try:
                print(f"\n处理第 {row_num} 行的URL: {url}")

                # 4. 打开网页
                print("正在打开网页...")
                driver.get(url)
                time.sleep(3)  # 等待页面加载

                # 5. 检查是否有"请点此进入"或类似提示
                try:
                    page_text = driver.page_source
                    if "请点此进入" in page_text or "请点击进入" in page_text or "点击进入" in page_text:
                        print("检测到需要点击进入的提示")

                        # 尝试找到并点击进入链接
                        entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "进入")
                        if not entry_links:
                            entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "点此")
                        if not entry_links:
                            entry_links = driver.find_elements(By.PARTIAL_LINK_TEXT, "点击")

                        if entry_links:
                            print(f"找到 {len(entry_links)} 个可能的进入链接")
                            entry_links[0].click()
                            print("已点击进入链接")
                            time.sleep(3)  # 等待新页面加载
                        else:
                            print("未找到明确的进入链接，尝试查找按钮")
                            buttons = driver.find_elements(By.TAG_NAME, "button")
                            for button in buttons:
                                if "进入" in button.text or "点此" in button.text or "点击" in button.text:
                                    button.click()
                                    print("已点击进入按钮")
                                    time.sleep(3)
                                    break
                except Exception as e:
                    print(f"检查进入提示时出错: {e}")

                # 6. 获取页面源代码
                page_source = driver.page_source

                # 7. 使用正则表达式查找magnet链接
                magnet_pattern = r'magnet:\?xt=urn:btih:[a-zA-Z0-9]{40}.*?(?=\'|"|&|<|\\s)'
                magnet_links = re.findall(magnet_pattern, page_source)

                if not magnet_links:
                    # 尝试更宽松的匹配模式
                    magnet_pattern = r'magnet:\?xt=[^\'"\s<>]+'
                    magnet_links = re.findall(magnet_pattern, page_source)

                if magnet_links:
                    # 取第一个找到的magnet链接
                    magnet_link = magnet_links[0]
                    print(f"找到magnet链接: {magnet_link}")

                    # 检查是否有重复的磁力链接（检查当前行之前的所有C列值）
                    is_duplicate = False
                    for check_row in range(2, row_num):
                        existing_link = sheet[f'C{check_row}'].value
                        if existing_link and existing_link == magnet_link:
                            is_duplicate = True
                            print(f"发现重复的磁力链接，位于C{check_row}")
                            break

                    # 8. 将magnet链接写入C列
                    sheet[f'C{row_num}'] = magnet_link

                    # 在D列标记状态
                    if is_duplicate:
                        sheet[f'D{row_num}'] = "重复"
                        print(f"重复磁力链接，已在D{row_num}单元格标记")
                    else:
                        sheet[f'D{row_num}'] = "成功"
                        print(f"磁力链接已写入C{row_num}单元格")

                    processed_count += 1

                else:
                    print("在页面中没有找到magnet链接")
                    # 在C列留空，D列标记"未找到磁力链接"
                    sheet[f'C{row_num}'] = None
                    sheet[f'D{row_num}'] = "未找到磁力链接"
                    print(f"已在D{row_num}单元格标记'未找到磁力链接'")

                    # 创建调试目录并保存页面源代码
                    debug_dir = "NoFind"
                    if not os.path.exists(debug_dir):
                        os.makedirs(debug_dir)

                    # 使用行号作为文件名
                    filename = f"row_{row_num}_debug.html"
                    debug_file = os.path.join(debug_dir, filename)

                    with open(debug_file, "w", encoding="utf-8") as f:
                        f.write(page_source)
                    print(f"已保存页面源代码到 {debug_file} 以供调试")

                # 每处理5个URL保存一次，防止数据丢失
                if processed_count % 5 == 0:
                    wb.save(excel_file)
                    print(f"已处理 {processed_count} 个URL，自动保存Excel文件")

            except Exception as e:
                print(f"处理URL {url} 时发生错误: {e}")
                # 在C列留空，D列标记错误信息
                sheet[f'C{row_num}'] = None
                sheet[f'D{row_num}'] = f"错误: {str(e)[:30]}..."  # 截断错误信息避免过长
                # 继续处理下一个URL
                continue

        # 处理完成后，再次全面检查重复项
        print("\n处理完成，开始全面检查重复项...")
        magnet_values = {}
        for row in range(2, sheet.max_row + 1):
            magnet_value = sheet[f'C{row}'].value
            if magnet_value and str(magnet_value).strip() != '':
                if magnet_value in magnet_values:
                    # 标记当前行和之前的所有相同行为重复
                    sheet[f'D{row}'] = "重复"
                    sheet[f'D{magnet_values[magnet_value]}'] = "重复"
                    print(f"标记重复: 行 {magnet_values[magnet_value]} 和 行 {row}")
                else:
                    magnet_values[magnet_value] = row

        # 最终保存Excel文件
        wb.save(excel_file)
        print(f"\n所有URL处理完成，共处理了 {processed_count} 个URL，Excel文件已保存")
        print(f"处理结果：")
        print(f"- C列：磁力链接（如果找到的话）")
        print(f"- D列：状态标记（成功/重复/未找到磁力链接/错误）")

    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")

    finally:
        # 关闭浏览器
        if 'driver' in locals():
            driver.quit()


# 运行函数
if __name__ == "__main__":
    extract_magnet_links_from_excel()